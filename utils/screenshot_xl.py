from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta
import os
import sys
from dotenv import load_dotenv
load_dotenv()
import time
from utils.get_driver import get_driver
from .get_driver import get_driver
import openpyxl
from PIL import Image
import io


def screenshot_xl(startDate, endDate):
    """
    Take continuous screenshots of week data from start date to end date.
    Captures Code Time (CT) and Active Code Time (ACT) for each date.
    
    Args:
        startDate (str): Start date in format 'YYYY-MM-DD'
        endDate (str): End date in format 'YYYY-MM-DD'
    """
    
    driver = get_driver()
    
    # Parse dates
    start = datetime.strptime(startDate, "%Y-%m-%d")
    end = datetime.strptime(endDate, "%Y-%m-%d")
    
    # Setup Excel file
    xl_path = os.path.expanduser("~/TrackCoder/weekly_screenshots_data.xlsx")
    os.makedirs(os.path.dirname(xl_path), exist_ok=True)
    
    # Create or load workbook with error handling
    existing_dates = []
    try:
        if os.path.exists(xl_path):
            try:
                workbook = openpyxl.load_workbook(xl_path)
                sheet = workbook.active
                # Get existing dates to avoid duplicates
                date_column = sheet['A']
                existing_dates = [cell.value for cell in date_column if cell.value is not None and cell.value != "Date"]
            except Exception as e:
                print(f"⚠️  Warning: Could not load existing Excel file ({e}). Creating a new one.")
                # Delete corrupted file and create new one
                os.remove(xl_path)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["Date", "Day", "Code Time (CT)", "Active Code Time (ACT)", "Screenshot Path"])
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Date", "Day", "Code Time (CT)", "Active Code Time (ACT)", "Screenshot Path"])
    except Exception as e:
        print(f"❌ Error setting up Excel file: {e}")
        return
    
    try:
        print("\n" + "=" * 70)
        print(f"📸 Starting Weekly Screenshot Capture")
        print(f"📅 Date Range: {startDate} to {endDate}")
        print("=" * 70 + "\n")
        
        # Get list of weeks (using Monday as week start)
        current_date = start
        processed_count = 0
        skipped_count = 0
        
        # Find the Monday of the start week
        days_to_monday = current_date.weekday()  # 0 = Monday, 6 = Sunday
        week_start = current_date - timedelta(days=days_to_monday)
        
        while week_start <= end:
            week_str = week_start.strftime("%Y-%m-%d")
            week_end = week_start + timedelta(days=6)
            
            print(f"\n🔄 Processing Week: {week_str} to {week_end.strftime('%Y-%m-%d')}")
            
            # Check if week already exists
            if week_str in existing_dates:
                print(f"⏭️  Skipped - Data already exists for week {week_str}")
                skipped_count += 1
                week_start += timedelta(days=7)
                continue
            
            try:
                # Navigate to the specific week (use Monday as week_of parameter)
                driver.get(f"https://app.software.com/code_time?week_of={week_str}")
                
                # Wait for login page or dashboard
                try:
                    WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.NAME, "email"))
                    )
                    # Login is needed
                    userName_field = driver.find_element(By.ID, "email")
                    password_field = driver.find_element(By.ID, "password")
                    signIn_btn = driver.find_element(By.CLASS_NAME, "btn-primary")
                    
                    userName_field.send_keys(os.getenv("userEmail_code_time"))
                    password_field.send_keys(os.getenv("password_code_time"))
                    signIn_btn.click()
                    
                    print("✅ Logged in successfully")
                except:
                    # Already logged in
                    pass
                
                # Wait for the graph to load
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "highcharts-background"))
                )
                
                # Additional wait for data to fully load
                time.sleep(3)
                
                # Calculate weekly totals by summing all days
                total_act_minutes = 0
                total_ct_minutes = 0
                
                days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
                
                for day_abbr in days:
                    try:
                        # Extract Active Code Time (ACT) for each day
                        xpath_active = f'//*[contains(@aria-label, "{day_abbr},") and contains(@aria-label, "Active Code Time")]'
                        element_active = driver.find_element(By.XPATH, xpath_active)
                        aria_label_active = element_active.get_attribute('aria-label')
                        
                        site_act = aria_label_active.split(",")[1].strip().split()[0]
                        active_code_time = round(float(site_act.rstrip('.')), 2)
                        total_act_minutes += active_code_time
                        
                        # Extract Code Time (CT) for each day
                        xpath_code_time = f'//*[contains(@aria-label, "{day_abbr},") and contains(@aria-label, "Code Time")]'
                        element_code_time = driver.find_element(By.XPATH, xpath_code_time)
                        aria_label_code_time = element_code_time.get_attribute('aria-label')
                        
                        site_ct = aria_label_code_time.split(",")[1].strip().split()[0]
                        code_time = round(float(site_ct.rstrip('.')), 2)
                        total_ct_minutes += code_time
                        
                    except Exception as e:
                        # Day might not have data, skip it
                        pass
                
                # Convert totals to hours:minutes format
                hours_active = int(total_act_minutes // 60)
                minutes_active = int(total_act_minutes % 60)
                ACT = f"{hours_active:02}:{minutes_active:02}"
                
                ct_hours = int(total_ct_minutes // 60)
                minutes_ct = int(total_ct_minutes % 60)
                
                # Calculate total CT (Active Code Time + Code Time)
                total_minutes = total_act_minutes + total_ct_minutes
                total_hours = int(total_minutes // 60)
                remaining_minutes = int(total_minutes % 60)
                CT = f"{total_hours:02}:{remaining_minutes:02}"
                
                # Take screenshot
                screenshot_dir = os.path.expanduser("~/TrackCoder/weekly_screenshots/")
                os.makedirs(screenshot_dir, exist_ok=True)
                
                # Scroll to top
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(1)
                
                # Find and scroll to graph element
                graph = driver.find_element(By.CLASS_NAME, "highcharts-background")
                driver.execute_script("arguments[0].scrollIntoView();", graph)
                time.sleep(1)
                
                # Take screenshot with week date in filename
                screenshot_path = os.path.join(screenshot_dir, f"week_{week_str}.png")
                driver.save_screenshot(screenshot_path)
                
                print(f"   📸 Screenshot saved: {screenshot_path}")
                print(f"   ⏱️  Weekly CT: {CT} | ACT: {ACT}")
                
                # Add data to Excel
                sheet.append([week_str, f"Week of {week_start.strftime('%b %d')}", CT, ACT, screenshot_path])
                workbook.save(xl_path)
                
                processed_count += 1
                print(f"   ✅ Data saved to Excel")
                
            except Exception as e:
                print(f"   ❌ Error processing week {week_str}: {e}")
                # Add empty entry to mark the attempt
                sheet.append([week_str, f"Week of {week_start.strftime('%b %d')}", "Error", "Error", f"Error: {str(e)}"])
                workbook.save(xl_path)
            
            # Move to next week
            week_start += timedelta(days=7)
        
        print("\n" + "=" * 70)
        print(f"✅ Weekly Screenshot Capture Completed!")
        print(f"📊 Processed: {processed_count} weeks")
        print(f"⏭️  Skipped: {skipped_count} weeks (already existed)")
        print(f"📁 Excel file: {xl_path}")
        print(f"📸 Screenshots: {os.path.expanduser('~/TrackCoder/weekly_screenshots/')}")
        print("=" * 70 + "\n")
        
    except Exception as e:
        print(f"\n❌ Fatal error in screenshot_xl: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Save and close workbook properly
        try:
            workbook.save(xl_path)
            workbook.close()
        except:
            pass
        
        driver.quit()
        print("🔒 Browser closed")



